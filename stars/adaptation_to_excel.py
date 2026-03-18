"""
DMS Site Adaptation Report → Excel Converter

Parses the plaintext DMS Site Adaptation Report and writes each SUM section
to its own Excel tab with bolded headers. Multi-column-block tables (where
the same row numbers appear across several horizontal blocks) are merged
into a single wide table.

Requirements:
    pip install openpyxl

Usage:
    python dms_to_excel.py <input_file.txt> [output_file.xlsx]
"""

import re
import sys
from pathlib import Path
from collections import OrderedDict

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# 1.  Parsing helpers
# ---------------------------------------------------------------------------

SKIP_SECTIONS = {"3.2"}  # sections to ignore entirely

TABLE_NAME_RE = re.compile(r"^\*{5,}\s*(.+?)\s*\*{5,}\s*$", re.MULTILINE)


def parse_sections(text: str) -> OrderedDict:
    """
    Return {para_num: (title, body_text)} for every section in the file.

    Strategy: find every line containing 'SUM Paragraph Number:' and then
    work outward to locate the enclosing banner and the body that follows.
    """
    lines = text.splitlines()
    sections = OrderedDict()

    # First pass: find all banner regions.
    # A banner is a block of lines between two full-width ### separator lines.
    # We look for lines that are all '#' (with optional whitespace).
    sep_indices = [
        i for i, ln in enumerate(lines)
        if re.match(r"^\s*#{20,}\s*$", ln)
    ]

    # Group consecutive separator lines into "separator blocks"
    # Then identify banners: a separator block, content lines, separator block.
    # We pair them up to find banner regions.
    sep_blocks: list[tuple[int, int]] = []  # (start_idx, end_idx) inclusive
    if sep_indices:
        block_start = sep_indices[0]
        prev = sep_indices[0]
        for si in sep_indices[1:]:
            if si == prev + 1:
                prev = si
            else:
                sep_blocks.append((block_start, prev))
                block_start = si
                prev = si
        sep_blocks.append((block_start, prev))

    # A banner = sep_block + content lines containing title/SUM + sep_block
    # We find banners by looking for pairs of consecutive sep_blocks with
    # content lines between them that contain 'SUM Paragraph Number'.
    banners: list[dict] = []  # {start, end, title, para}
    for bi in range(len(sep_blocks) - 1):
        _, end1 = sep_blocks[bi]
        start2, _ = sep_blocks[bi + 1]
        if start2 <= end1 + 1:
            continue  # no content lines between

        content_lines = lines[end1 + 1: start2]
        content = "\n".join(content_lines)

        para_match = re.search(r"SUM Paragraph Number:\s*([\d.]+)", content)
        if not para_match:
            continue

        para = para_match.group(1).strip()

        # Extract title: look for lines with ## ... content ... ##
        # that are NOT the SUM line and NOT the Report Type etc.
        title = ""
        for cl in content_lines:
            stripped = cl.strip().strip("#").strip()
            if not stripped:
                continue
            if "SUM Paragraph Number" in stripped:
                continue
            # This should be the title line
            title = stripped
            break

        banners.append({
            "start": sep_blocks[bi][0],
            "end": sep_blocks[bi + 1][1],
            "title": title,
            "para": para,
        })

    # Now extract body text between banners
    for bi, banner in enumerate(banners):
        para = banner["para"]
        title = banner["title"]

        if para in SKIP_SECTIONS:
            continue

        body_start = banner["end"] + 1
        body_end = banners[bi + 1]["start"] if bi + 1 < len(banners) else len(lines)
        body = "\n".join(lines[body_start:body_end])

        if para in sections:
            sections[para] = (sections[para][0], sections[para][1] + "\n" + body)
        else:
            sections[para] = (title, body)

    return sections


def _parse_row(line: str) -> list[str]:
    """Split a pipe-delimited row into cell values, stripping whitespace."""
    parts = line.strip().strip("|").split("|")
    return [p.strip() for p in parts]


def _is_data_line(line: str) -> bool:
    """True if the line is a pipe-delimited data row (not a separator/note)."""
    stripped = line.strip()
    if not stripped.startswith("|"):
        return False
    if _is_separator(stripped):
        return False
    if _is_note_line(stripped):
        return False
    return True


def _is_separator(line: str) -> bool:
    stripped = line.strip()
    return bool(re.match(r"^[|\-=+\s]+$", stripped)) and "|" in stripped and not re.search(r"[a-zA-Z0-9]", stripped)


def _is_note_line(line: str) -> bool:
    """Lines like | Note: ... | that span the full width."""
    stripped = line.strip()
    if not stripped.startswith("|"):
        return False
    inner = stripped.strip("|").strip()
    return inner.startswith("Note:")


def parse_tables(body: str) -> list[tuple[str, list[list[str]]]]:
    """
    Extract every named table from a section body.
    Returns [(table_name, rows), ...] where rows[0] is the header.
    Multi-block tables (same row # column repeated) are merged horizontally.
    """
    tables: list[tuple[str, list[list[str]]]] = []
    names = list(TABLE_NAME_RE.finditer(body))

    for idx, nm in enumerate(names):
        tname = nm.group(1).strip()
        start = nm.end()
        end = names[idx + 1].start() if idx + 1 < len(names) else len(body)
        chunk = body[start:end]

        lines = chunk.splitlines()

        # Collect contiguous "blocks" — each block is a header + data rows
        blocks: list[list[list[str]]] = []
        current_block: list[list[str]] = []

        for line in lines:
            stripped = line.strip()
            if not stripped:
                if current_block:
                    blocks.append(current_block)
                    current_block = []
                continue

            if _is_separator(stripped):
                continue

            if _is_note_line(stripped):
                continue

            if _is_data_line(line):
                row = _parse_row(line)
                current_block.append(row)

        if current_block:
            blocks.append(current_block)

        if not blocks:
            continue

        merged = _merge_blocks(blocks)
        tables.append((tname, merged))

    return tables


def _merge_blocks(blocks: list[list[list[str]]]) -> list[list[str]]:
    """
    Merge multiple column-blocks into one wide table.
    Blocks that share the same '#.' row-number column get merged horizontally.
    """
    if len(blocks) <= 1:
        return blocks[0] if blocks else []

    groups: list[list[list[list[str]]]] = []
    current_group: list[list[list[str]]] = [blocks[0]]

    for blk in blocks[1:]:
        prev = current_group[0]
        same_header = (
            blk[0][0] == prev[0][0] if (blk and prev and blk[0] and prev[0]) else False
        )
        same_row_count = len(blk) == len(prev)
        if same_header and same_row_count:
            current_group.append(blk)
        else:
            groups.append(current_group)
            current_group = [blk]
    groups.append(current_group)

    result: list[list[str]] = []
    for gi, group in enumerate(groups):
        if gi > 0:
            result.append([])  # blank separator row

        if len(group) == 1:
            result.extend(group[0])
        else:
            base = group[0]
            for row_idx in range(len(base)):
                merged_row = list(base[row_idx])
                for extra_blk in group[1:]:
                    merged_row.extend(extra_blk[row_idx][1:])
                result.append(merged_row)

    return result


# ---------------------------------------------------------------------------
# 2.  Excel writing
# ---------------------------------------------------------------------------

BOLD = Font(bold=True)


def _sanitize_sheet_name(name: str) -> str:
    """Excel sheet names: max 31 chars, no special chars."""
    name = re.sub(r"[\\/*?\[\]:]", "", name)
    return name[:31]


def write_workbook(sections: OrderedDict, output_path: str):
    wb = Workbook()
    # Remove default sheet
    default_sheet = wb.active

    sheets_created = 0
    for para, (title, body) in sections.items():
        sheet_name = _sanitize_sheet_name(f"{para} - {title}")
        ws = wb.create_sheet(title=sheet_name)
        sheets_created += 1

        tables = parse_tables(body)
        if not tables:
            # Still keep the sheet, just leave it empty
            continue

        row_cursor = 1
        for tidx, (tname, rows) in enumerate(tables):
            if tidx > 0:
                row_cursor += 2  # blank rows between tables

            # Write table name as a bold label
            cell = ws.cell(row=row_cursor, column=1, value=_clean_string(tname))
            cell.font = BOLD
            row_cursor += 1

            if not rows:
                continue

            # First row is the header — write it bold
            for ci, val in enumerate(rows[0], start=1):
                cell = ws.cell(row=row_cursor, column=ci, value=_clean_string(val))
                cell.font = BOLD
            row_cursor += 1

            # Data rows
            for row in rows[1:]:
                if not any(row):  # blank separator
                    row_cursor += 1
                    continue
                for ci, val in enumerate(row, start=1):
                    ws.cell(row=row_cursor, column=ci, value=_auto_type(val))
                row_cursor += 1

        # Auto-fit column widths (approximate)
        for col_cells in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 3, 50)

    # Remove the default empty sheet if we created others
    if sheets_created > 0 and default_sheet is not None:
        wb.remove(default_sheet)
    elif sheets_created == 0:
        # No sections found — leave default sheet with a message
        default_sheet.cell(row=1, column=1, value="No sections found in input file.")
        print("WARNING: No sections were found. Check the input file format.")

    wb.save(output_path)
    print(f"Saved workbook to: {output_path}")


def _clean_string(val: str) -> str:
    """Remove illegal XML/Excel characters (control chars except tab/newline)."""
    # openpyxl rejects control characters in the range \x00-\x08, \x0b-\x0c, \x0e-\x1f
    return re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", "", val)


def _auto_type(val: str):
    """Convert string to int/float if possible, otherwise return as-is."""
    if not val:
        return val
    try:
        return int(val)
    except ValueError:
        pass
    try:
        return float(val)
    except ValueError:
        pass
    return _clean_string(val)


# ---------------------------------------------------------------------------
# 3.  Main
# ---------------------------------------------------------------------------

def main():
    if len(sys.argv) < 2:
        print("Usage: python dms_to_excel.py <input_file.txt> [output_file.xlsx]")
        sys.exit(1)

    input_path = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) > 2 else str(
        Path(input_path).with_suffix(".xlsx")
    )

    text = Path(input_path).read_text(encoding="utf-8", errors="replace")

    # Debug: show first few lines to verify file loaded
    file_lines = text.splitlines()
    print(f"File loaded: {len(file_lines)} lines, {len(text)} chars")

    # Debug: count ### separator lines
    sep_count = sum(1 for ln in file_lines if re.match(r"^\s*#{20,}\s*$", ln))
    print(f"Found {sep_count} separator lines (###...)")

    # Debug: count SUM paragraph references
    sum_count = len(re.findall(r"SUM Paragraph Number:", text))
    print(f"Found {sum_count} SUM Paragraph Number references")

    sections = parse_sections(text)

    print(f"Parsed {len(sections)} sections:")
    for para, (title, _) in sections.items():
        print(f"  {para} - {title}")

    write_workbook(sections, output_path)


if __name__ == "__main__":
    main()
